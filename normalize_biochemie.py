"""
Normalizace dat z laboratorní příručky Biochemie.
Transponovaná Excel tabulka → CSV + JSON + SQLite (1 řádek = 1 test).
"""

import csv
import json
import sqlite3
import sys
from pathlib import Path

import openpyxl

SOURCE = Path("STABILITA_tabulky/Export_laboratorni_prirucky_z_webu_Biochemie_15012026.xlsx")
OUT_DIR = Path("STABILITA_data")

LABEL_MAP = {
    "Zkratka":                          "zkratka",
    "Metoda":                           "nazev",
    "OpenLims ID":                      "openlims_id",
    "BHSI | Princip stanovení:":        "princip_stanoveni",
    "BHSI | Statim:":                   "statim",
    "BHSI | Vyšetřovaný materiál:":     "material",
    "BHSI | Primární materiál:":        "primarni_material",
    "BHSI | Odebíraný materiál - typ:": "typ_materialu",
    "BHSI | Stabilita:":                "stabilita",
    "BHSI | Jednotka:":                 "jednotka",
    "BHSI | Frekvence stanovení:":      "frekvence",
    "BHSI | Doba odezvy:":              "doba_odezvy",
    "BHSI | Interpretace:":             "interpretace",
    "BHSI | Klinické informace:":       "klinicke_informace",
    "BHSI | Poznámka:":                 "poznamka",
    "ALL | Výkon:":                     "vzp_vykon",
    "Prováděno v:":                     "provadeno_v",
    "Kontrola - zadáno:":               "kontrola_zadano",
    "Kontrola - VLEK:":                 "kontrola_vlek",
    "Kontrola - VA:":                   "kontrola_va",
    "Kontrola - SMK:":                  "kontrola_smk",
}

FIELDS = [
    "oddeleni", "kategorie", "zkratka", "nazev", "openlims_id",
    "princip_stanoveni", "statim", "material", "primarni_material",
    "typ_materialu", "stabilita", "jednotka", "frekvence", "doba_odezvy",
    "interpretace", "klinicke_informace", "poznamka",
    "vzp_vykon", "provadeno_v",
    "kontrola_zadano", "kontrola_vlek", "kontrola_va", "kontrola_smk",
]


def cell_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_sheet(ws, sheet_name):
    rows = [list(row) for row in ws.iter_rows(values_only=True)]

    # Najdi radek a sloupec s "Zkratka"
    label_col = None
    zkratka_row = None
    for ri, row in enumerate(rows):
        for ci, cell in enumerate(row):
            if cell_str(cell) == "Zkratka":
                label_col = ci
                zkratka_row = ri
                break
        if label_col is not None:
            break

    if label_col is None:
        print(f"  WARN: list '{sheet_name}' – nenalezen řádek Zkratka, přeskočeno", file=sys.stderr)
        return []

    data_start_col = label_col + 2

    # Sestav label → row_index mapping
    label_to_row = {}
    for ri, row in enumerate(rows):
        label = cell_str(row[label_col]) if label_col < len(row) else None
        if label:
            label_to_row[label] = ri

    # Zjisti šířku dat
    max_col = max(len(row) for row in rows)

    records = []
    for col in range(data_start_col, max_col):
        zkratka_val = cell_str(rows[zkratka_row][col]) if col < len(rows[zkratka_row]) else None
        if not zkratka_val:
            continue

        record = {"oddeleni": "Biochemie", "kategorie": sheet_name}
        for label_raw, field_name in LABEL_MAP.items():
            ri = label_to_row.get(label_raw)
            if ri is not None and col < len(rows[ri]):
                record[field_name] = cell_str(rows[ri][col])
            else:
                record[field_name] = None

        records.append(record)

    return records


def main():
    OUT_DIR.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    all_records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        recs = parse_sheet(ws, sheet_name)
        print(f"  {sheet_name}: {len(recs)} testů")
        all_records.extend(recs)

    wb.close()
    print(f"\nCelkem: {len(all_records)} testů\n")

    # --- CSV (UTF-8 BOM pro Excel) ---
    csv_path = OUT_DIR / "biochemie.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS, delimiter=";", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_records)
    print(f"CSV:    {csv_path}")

    # --- JSON ---
    json_path = OUT_DIR / "biochemie.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2)
    print(f"JSON:   {json_path}")

    # --- SQLite ---
    db_path = OUT_DIR / "lab_catalog.db"
    con = sqlite3.connect(db_path)
    cur = con.cursor()
    cur.execute("DROP TABLE IF EXISTS testy")
    col_defs = ", ".join(
        f"{f} INTEGER" if f == "openlims_id" else f"{f} TEXT"
        for f in FIELDS
    )
    cur.execute(f"CREATE TABLE testy ({col_defs})")
    for rec in all_records:
        vals = []
        for f in FIELDS:
            v = rec.get(f)
            if f == "openlims_id" and v is not None:
                try:
                    v = int(v)
                except (ValueError, TypeError):
                    v = None
            vals.append(v)
        cur.execute(f"INSERT INTO testy VALUES ({','.join(['?']*len(FIELDS))})", vals)
    con.commit()

    # Ověření
    print(f"SQLite: {db_path}")
    print("\nPočty testů per kategorie:")
    for row in cur.execute("SELECT kategorie, COUNT(*) AS n FROM testy GROUP BY kategorie ORDER BY n DESC"):
        print(f"  {row[0]}: {row[1]}")
    print(f"\nTesty bez zkratky:  {cur.execute('SELECT COUNT(*) FROM testy WHERE zkratka IS NULL').fetchone()[0]}")
    print(f"Testy bez názvu:    {cur.execute('SELECT COUNT(*) FROM testy WHERE nazev IS NULL').fetchone()[0]}")
    statim_count = cur.execute("SELECT COUNT(*) FROM testy WHERE statim = 'Yes'").fetchone()[0]
    print(f"STATIM testy:       {statim_count}")
    con.close()


if __name__ == "__main__":
    main()
