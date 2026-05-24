"""
Normalizace všech laboratorních příruček STABILITA → SQLite + CSV.
Zpracovává: Biochemie, Hematologie, Moče, Sérologie, Bakteriologie, PCR.
Každý běh přepíše databázi (DROP TABLE), aby nevznikaly duplikáty.
"""

import csv
import sqlite3
from pathlib import Path

import openpyxl

OUT_DIR = Path("STABILITA_data")

# ---------------------------------------------------------------------------
# Mapování labelů → názvy polí (Typ A — BHSI: Biochemie, Hematologie, Moče, Sérologie)
# ---------------------------------------------------------------------------
BHSI_MAP = {
    "Zkratka":                          "zkratka",
    "Metoda":                           "nazev",
    "OpenLims ID":                      "openlims_id",
    "BHSI | Princip stanovení:":        "princip_stanoveni",
    "BHSI | Statim:":                   "statim",
    "BHSI | Vyšetřovaný materiál:":     "material",
    "BHSI | Primární materiál:":        "primarni_material",
    "BHSI | Odebíraný materiál - typ:": "typ_materialu",
    "BHSI | Stabilita:":                "stabilita",
    "BHSI | Jednotka:":                 "jednotka",
    "BHSI | Frekvence stanovení:":      "frekvence",
    "BHSI | Doba odezvy:":              "doba_odezvy",
    "BHSI | Interpretace:":             "interpretace",
    "BHSI | Klinické informace:":       "klinicke_informace",
    "BHSI | Poznámka:":                 "poznamka",
    "ALL | Výkon:":                     "vzp_vykon",
    "Prováděno v:":                     "provadeno_v",
    "Kontrola - zadáno:":               "kontrola_zadano",
    "Kontrola - VLEK:":                 "kontrola_vlek",
    "Kontrola - VA:":                   "kontrola_va",
    "Kontrola - SMK:":                  "kontrola_smk",
}

# ---------------------------------------------------------------------------
# Mapování labelů → názvy polí (Typ B/C — Mik: Bakteriologie + PCR)
# ---------------------------------------------------------------------------
MIK_MAP = {
    "Zkratka":                                  "zkratka",
    "Název":                                    "nazev",
    "BHSI | Statim:":                           "statim",
    "Mik | Typ výtěrovky:":                     "typ_materialu",
    "Mik | Postup odběru:":                     "klinicke_informace",
    "Mik PCR | Uchování:":                      "stabilita",
    "Mik PCR | Transport:":                     "material",
    "Mik PCR | Poznámka:":                      "poznamka",
    "Mik PCR | Doba odezvy:":                   "doba_odezvy",
    # PCR-specifické (Typ C) — přebijí Mik hodnoty tam, kde existují
    "PCR | Abstrakt:":                          "interpretace",
    "PCR | Typ vyšetřovaného materiálu:":       "primarni_material",
    "PCR | Zkumavka / odběrovka:":              "typ_materialu",
    # Ostatní
    "Typ odběrového materiálu":                 "primarni_material",
    "Prováděno v:":                             "provadeno_v",
    "Kontrola - zadáno:":                       "kontrola_zadano",
    "Kontrola - VLEK:":                         "kontrola_vlek",
    "Kontrola - VA:":                           "kontrola_va",
    "Kontrola - SMK:":                          "kontrola_smk",
}

# ---------------------------------------------------------------------------
# Všechny výstupní pole (pořadí sloupců v DB/CSV)
# ---------------------------------------------------------------------------
FIELDS = [
    "oddeleni", "kategorie", "zkratka", "nazev", "openlims_id",
    "princip_stanoveni", "statim", "material", "primarni_material",
    "typ_materialu", "stabilita", "jednotka", "frekvence", "doba_odezvy",
    "interpretace", "klinicke_informace", "poznamka",
    "vzp_vykon", "provadeno_v",
    "kontrola_zadano", "kontrola_vlek", "kontrola_va", "kontrola_smk",
]

# ---------------------------------------------------------------------------
# Seznam souborů: (cesta, oddělení, listy k vynechání)
# ---------------------------------------------------------------------------
FILE_LIST = [
    (
        "STABILITA_tabulky/Export_laboratorni_prirucky_z_webu_Biochemie_15012026.xlsx",
        "Biochemie",
        [],
    ),
    (
        "STABILITA_tabulky/Export_laboratorni_prirucky_z_webu_Hematologie_15102025.xlsx",
        "Hematologie",
        [],
    ),
    (
        "STABILITA_tabulky/Export_laboratorni_prirucky_z_webu_Moce_19112021.xlsx",
        "Moče",
        [],
    ),
    (
        "STABILITA_tabulky/Export_laboratorni_prirucky_z_webu_Serologie_30012026.xlsx",
        "Sérologie",
        [],
    ),
    (
        "STABILITA_tabulky/Export_laboratorni_prirucky_z_webu_Bakteriologie_09052024.xlsx",
        "Bakteriologie",
        ["Odběrový materiál - zkratky"],  # referenční tabulka, ne seznam testů
    ),
    (
        "STABILITA_tabulky/Export_laboratorni_prirucky_z_webu_PCR_31072023.xlsx",
        "PCR",
        [],
    ),
]


def cell_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def detect_label_map(label_to_row: dict) -> dict:
    """Vrátí správnou label mapu podle přítomnosti BHSI nebo Mik labelů."""
    if "BHSI | Stabilita:" in label_to_row:
        return BHSI_MAP
    return MIK_MAP


def parse_sheet(ws, sheet_name: str, oddeleni: str) -> list[dict]:
    rows = [list(row) for row in ws.iter_rows(values_only=True)]

    # Dynamicky najdi řádek a sloupec s "Zkratka"
    label_col = None
    zkratka_row_idx = None
    for ri, row in enumerate(rows):
        for ci, cell in enumerate(row):
            if cell_str(cell) == "Zkratka":
                label_col = ci
                zkratka_row_idx = ri
                break
        if label_col is not None:
            break

    if label_col is None:
        print(f"  WARN [{sheet_name}]: řádek Zkratka nenalezen — přeskočeno")
        return []

    data_start_col = label_col + 2

    # Sestav label → row_index mapping
    label_to_row = {}
    for ri, row in enumerate(rows):
        label = cell_str(row[label_col]) if label_col < len(row) else None
        if label:
            label_to_row[label] = ri

    label_map = detect_label_map(label_to_row)
    max_col = max(len(row) for row in rows)

    records = []
    for col in range(data_start_col, max_col):
        zkratka_val = (
            cell_str(rows[zkratka_row_idx][col])
            if col < len(rows[zkratka_row_idx])
            else None
        )
        if not zkratka_val:
            continue

        record: dict = {"oddeleni": oddeleni, "kategorie": sheet_name}
        for label_raw, field_name in label_map.items():
            ri = label_to_row.get(label_raw)
            val = None
            if ri is not None and col < len(rows[ri]):
                val = cell_str(rows[ri][col])
            # Nepřepisovat existující hodnotu prázdnou (pro PCR Typ C — typ_materialu)
            if val is not None or field_name not in record:
                record[field_name] = val

        records.append(record)

    return records


def main():
    OUT_DIR.mkdir(exist_ok=True)

    # Připrav SQLite — vždy začni čistě (DROP TABLE zabraňuje duplikátům při re-run)
    db_path = OUT_DIR / "lab_catalog.db"
    con = sqlite3.connect(db_path)
    cur = con.cursor()
    cur.execute("DROP TABLE IF EXISTS testy")
    col_defs = ", ".join(
        "openlims_id INTEGER" if f == "openlims_id" else f"{f} TEXT"
        for f in FIELDS
    )
    cur.execute(f"CREATE TABLE testy ({col_defs})")

    grand_total = 0

    for path_str, oddeleni, exclude_sheets in FILE_LIST:
        path = Path(path_str)
        if not path.exists():
            print(f"CHYBA: soubor nenalezen: {path}")
            continue

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        dept_records: list[dict] = []

        print(f"\n=== {oddeleni} ===")
        for sheet_name in wb.sheetnames:
            if sheet_name in exclude_sheets:
                print(f"  [{sheet_name}]: přeskočeno (vyloučeno)")
                continue

            ws = wb[sheet_name]
            recs = parse_sheet(ws, sheet_name, oddeleni)
            print(f"  [{sheet_name}]: {len(recs)} testů")
            dept_records.extend(recs)

        wb.close()
        print(f"  Celkem {oddeleni}: {len(dept_records)} testů")
        grand_total += len(dept_records)

        # CSV per oddělení
        csv_name = oddeleni.lower().replace("č", "c").replace("é", "e") + ".csv"
        csv_path = OUT_DIR / csv_name
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(
                f, fieldnames=FIELDS, delimiter=";", extrasaction="ignore"
            )
            writer.writeheader()
            writer.writerows(dept_records)

        # Vložit do SQLite
        for rec in dept_records:
            vals = []
            for f in FIELDS:
                v = rec.get(f)
                if f == "openlims_id" and v is not None:
                    try:
                        v = int(v)
                    except (ValueError, TypeError):
                        v = None
                vals.append(v)
            cur.execute(
                f"INSERT INTO testy VALUES ({','.join(['?'] * len(FIELDS))})", vals
            )

    con.commit()

    # Souhrn
    print(f"\n{'='*50}")
    print(f"CELKEM testů: {grand_total}")
    print("\nPočty per oddělení:")
    for row in cur.execute(
        "SELECT oddeleni, COUNT(*) AS n FROM testy GROUP BY oddeleni ORDER BY n DESC"
    ):
        print(f"  {row[0]}: {row[1]}")

    print(f"\nTesty bez zkratky:  {cur.execute('SELECT COUNT(*) FROM testy WHERE zkratka IS NULL').fetchone()[0]}")
    print(f"Testy bez názvu:    {cur.execute('SELECT COUNT(*) FROM testy WHERE nazev IS NULL').fetchone()[0]}")

    # Ověření vyloučení referenčního listu
    bakt_ref = cur.execute(
        "SELECT COUNT(*) FROM testy WHERE oddeleni='Bakteriologie' AND kategorie='Odběrový materiál - zkratky'"
    ).fetchone()[0]
    print(f"Bakteriologie ref. list: {bakt_ref} záznamů (očekáváno: 0)")

    con.close()
    print(f"\nSQLite: {db_path.resolve()}")


if __name__ == "__main__":
    main()
